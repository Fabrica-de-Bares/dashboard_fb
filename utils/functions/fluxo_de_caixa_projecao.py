import streamlit as st
import pandas as pd
from datetime import timedelta
from utils.queries_fluxo_de_caixa import *
from workalendar.america import Brazil
from utils.functions.general_functions import *
import openpyxl
import os


def highlight_total_row(row):
    if row['Data'] == 'Total':
        return ['background-color: #f0f2f6; color: black;'] * len(row)
    else:
        return [''] * len(row)


def seletor_data_fim_padrao(key):
  return st.date_input(
    "Data de Fim",
    value=datetime.datetime.today() + timedelta(days=7),
    key=key,
    format="DD/MM/YYYY",
    min_value=datetime.datetime.today() + timedelta(days=1),
    max_value=datetime.datetime.today() + timedelta(days=365)
  )


def convert_to_datetime(df, cols):
  for col in cols:
      if col in df.columns:
          df[col] = pd.to_datetime(df[col])
      else:
          print(f"Atenção: coluna '{col}' não encontrada no DataFrame")
  return df


def filtrar_data_fim(dataframe, data_fim, categoria):
  data_fim = pd.Timestamp(data_fim)
  dataframe.loc[:, categoria] = pd.to_datetime(dataframe[categoria], errors='coerce')
  dataframe_filtered = dataframe.loc[
    (dataframe[categoria] <= data_fim)
  ]
  return dataframe_filtered


def prolongar_projecao(df_projecao_zig, dias_prolongados=7):
  df_projecao_zig = df_projecao_zig.copy()
  df_projecao_zig['Data'] = pd.to_datetime(df_projecao_zig['Data'])
  novas_linhas = []

  for empresa, grupo in df_projecao_zig.groupby('Empresa'):
    for i in range(1, dias_prolongados + 1):
      df_temp = grupo.copy()
      df_temp['Data'] += pd.Timedelta(days=7)  # Avança 7 dias mantendo padrão de semana
      novas_linhas.append(df_temp)

  df_projecao_estendida = pd.concat([df_projecao_zig] + novas_linhas, ignore_index=True)
  df_projecao_estendida = df_projecao_estendida.sort_values(by=['Empresa', 'Data']).reset_index(drop=True)
  df_projecao_estendida.drop_duplicates(subset=['Data', 'Empresa'], keep='first', inplace=True)

  return df_projecao_estendida



def config_projecao_bares(seletor_status_despesa, df_saldos_bancarios, df_valor_liquido, df_projecao_zig, df_receitas_extraord_proj, df_receitas_eventos_proj, df_despesas_aprovadas, df_despesas_pagas, df_despesas_provisionadas):
  # Acresenta mais uma semana de dados
  df_projecao_zig = prolongar_projecao(df_projecao_zig, dias_prolongados=7)
  
  # Converter colunas de data
  dfs = [df_saldos_bancarios, 
         df_valor_liquido, 
         df_projecao_zig, 
         df_receitas_extraord_proj, 
         df_receitas_eventos_proj, 
         df_despesas_aprovadas, 
         df_despesas_pagas
        ]
  if seletor_status_despesa == 'Todas Previstas':
    dfs = dfs + [df_despesas_provisionadas]
  
  for df in dfs:
    if df is not None and not df.empty:
        df = convert_to_datetime(df, ['Data'])

  # Merge dataframes
  # manter apenas dfs válidos
  dfs_validos = [df for df in dfs if df is not None and not df.empty]

  # merge progressivo
  merged_df = dfs_validos[0]
  for df in dfs_validos[1:]:
      merged_df = pd.merge(
          merged_df,
          df,
          on=['Data', 'Empresa'],
          how='outer'
      )
  # Preenchendo valores nulos com 0 e renomeando colunas
  merged_df = merged_df.fillna(0)
  merged_df = merged_df.rename(columns={'Valor_Projetado': 'Valor_Projetado_Zig'})
 
  # Ajustando formatação
  cols = ['Saldo_Inicio_Dia', 'Valor_Liquido_Recebido', 'Valor_Projetado_Zig', 'Receita_Projetada_Extraord', 'Receita_Projetada_Eventos', 'Despesas_Aprovadas_Pendentes', 'Despesas_Pagas']
  if seletor_status_despesa == 'Todas Previstas':
    cols = cols + ['Provisões Pendentes']

  # Caso T_PROVISOES_PADRAO esteja vazia
  for col in cols:
    if col not in merged_df.columns:
      merged_df[col] = 0 # Zera a coluna de valor de Provisões Pendentes
  
  merged_df[cols] = merged_df[cols].astype(float).round(2)

  return merged_df


# Função que filtra por data, aplica o multiplicador e faz a soma do 'Total' nos df
def filtra_soma_saldo_final(seletor_status_despesa, merged_df, data_fim, multiplicador):
  # Filtrando por data de fim
  merged_df_filtrado = filtrar_data_fim(merged_df, data_fim, 'Data')
  merged_df_filtrado = merged_df_filtrado.copy()

  # Aplicando lógica de negócios
  merged_df_filtrado.loc[:, 'Valor_Projetado_Zig'] = merged_df_filtrado.apply(lambda row: 0 if row['Valor_Liquido_Recebido'] > 0 else row['Valor_Projetado_Zig'], axis=1)
  merged_df_filtrado.loc[:, 'Valor_Projetado_Zig'] = merged_df_filtrado['Valor_Projetado_Zig'] * multiplicador

  merged_df_filtrado.loc[:, 'Saldo_Final'] = (
     merged_df_filtrado['Saldo_Inicio_Dia'] + 
     merged_df_filtrado['Valor_Liquido_Recebido'] + 
     merged_df_filtrado['Valor_Projetado_Zig'] + 
     merged_df_filtrado['Receita_Projetada_Extraord'] + 
     merged_df_filtrado['Receita_Projetada_Eventos'] - 
     merged_df_filtrado['Despesas_Aprovadas_Pendentes'] - 
     merged_df_filtrado['Despesas_Pagas']
    )
  
  if seletor_status_despesa == 'Todas Previstas':
    merged_df_filtrado.loc[:, 'Saldo_Final'] = (
     merged_df_filtrado['Saldo_Final'] - 
     merged_df_filtrado['Provisões Pendentes']
    )
  merged_df_filtrado = df_format_date_brazilian(merged_df_filtrado, 'Data')

  return merged_df_filtrado


# Ordena df por data
def ordena_por_data(df):
  df['Data'] = pd.to_datetime(df['Data'], format="%d-%m-%Y", errors="coerce")
  if 'Casa' in df.columns:
    df = df.sort_values(by=["Casa", "Data"])
  else:
    df = df.sort_values(by=["Data"])
  
  # Reverte para string mantendo o Total no final
  df['Data'] = (df['Data'].dt.strftime("%d-%m-%Y").fillna("Total"))
  return df


def is_in_group(empresa, houses_to_group):
  return any(house in empresa for house in houses_to_group)


def config_grouped_projecao(df_projecao_bares, lojas_agrupadas):
  grouped_df = df_projecao_bares[df_projecao_bares['Empresa'].apply(lambda x: is_in_group(x, lojas_agrupadas))]
  grouped_df = grouped_df.groupby(['Data']).sum().reset_index()
  return grouped_df


# Filtra as despesas pendentes e pagas de acordo com a opção do seletor (aprovadas/previstas)
def filtra_categoria_despesas(df_despesas_aprovadas_previstas, seletor_status_despesa, tipo):
  hoje = pd.Timestamp.today().normalize()
  duas_semanas = hoje + pd.Timedelta(days=14)

  # 1) Filtra apenas despesas aprovadas pela diretoria
  if seletor_status_despesa == 'Apenas Aprovadas':
    df_categoria = df_despesas_aprovadas_previstas[
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 101) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 103)
    ].copy()

  # 2) Filtra pelas despesas previstas (aprovadas, pendentes, travados, e nulas)
  if seletor_status_despesa == 'Todas Previstas':
    df_categoria = df_despesas_aprovadas_previstas[
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 101) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 103) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 100) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 105) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'] == 108) |
      (df_despesas_aprovadas_previstas['Status_Diretoria'].isna())
    ].copy()
    
  df_categoria_agrupado = df_categoria.groupby(['Empresa', 'Previsao_Pgto'], as_index=False)['Valor_Liquido'].sum()

  # Filtra pela data de hoje até duas semanans a frente
  df_categoria_agrupado['Previsao_Pgto'] = pd.to_datetime(df_categoria_agrupado['Previsao_Pgto'], errors='coerce')

  df_categoria_agrupado = df_categoria_agrupado[
    (df_categoria_agrupado['Previsao_Pgto'] >= hoje) &
    (df_categoria_agrupado['Previsao_Pgto'] <= duas_semanas)
  ].copy()

  # Renomeia colunas
  if tipo == 'pendentes':
    df_categoria_agrupado = df_categoria_agrupado.rename(columns={'Previsao_Pgto':'Data', 'Valor_Liquido':'Despesas_Aprovadas_Pendentes'})

  if tipo == 'pagas':
    df_categoria_agrupado = df_categoria_agrupado.rename(columns={'Previsao_Pgto':'Data', 'Valor_Liquido':'Despesas_Pagas'})

  return df_categoria_agrupado


# Filtra o resultado da query de acordo com o seletor (aprovadas/previstas) e datas de inicio e fim
def filtra_detalhes_despesas(seletor_status_despesa, despesas_pendentes_pagas, data_inicio, data_fim):
  # Filtra de acordo com o seletor (aprovadas/previstas)
    if seletor_status_despesa == 'Apenas Aprovadas':
        df_despesas_pendentes_pagas = despesas_pendentes_pagas[
          (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 101) |
          (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 103)
        ].copy()
    else:
        df_despesas_pendentes_pagas = despesas_pendentes_pagas[
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 101) |
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 103) |
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 100) |
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 105) |
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'] == 108) |
            (despesas_pendentes_pagas['FK_Aprovacao_Diretoria'].isna())
        ].copy()

    # Filtra pelas datas de início e fim
    df_despesas_pendentes_pagas['Previsao_Pgto'] = pd.to_datetime(df_despesas_pendentes_pagas['Previsao_Pgto'], errors='coerce')
    df_despesas_pendentes_pagas = df_despesas_pendentes_pagas[
        ((df_despesas_pendentes_pagas['Previsao_Pgto'] >= data_inicio) &
        (df_despesas_pendentes_pagas['Previsao_Pgto'] <= data_fim))
    ]

    return df_despesas_pendentes_pagas


# Função que determina despesas proviões padrão não lançadas na T_DESPESA_RAPIDA
def despesas_provisionadas_lancadas(df_despesas_provisionadas, df_despesas_rapidas):
  df_despesas_provisionadas['Data Competência'] = pd.to_datetime(df_despesas_provisionadas['Data Competência'], errors='coerce').dt.normalize()
  df_despesas_provisionadas['Data Vencimento'] = pd.to_datetime(df_despesas_provisionadas['Data Vencimento'], errors='coerce').dt.normalize()
  df_despesas_provisionadas['Mes Competencia'] = df_despesas_provisionadas['Data Competência'].dt.month

  df_despesas_rapidas['Data_Competencia'] = pd.to_datetime(df_despesas_rapidas['Data_Competencia'], errors='coerce').dt.normalize()
  df_despesas_rapidas['Data_Vencimento'] = pd.to_datetime(df_despesas_rapidas['Data_Vencimento'], errors='coerce').dt.normalize()
  df_despesas_rapidas['Mes Competencia'] = df_despesas_rapidas['Data_Competencia'].dt.month

  df_merge = pd.merge( # Match entre T_PROVISOES_PADRAO e T_DESPESA_RAPIDA (casa, fornecedor, mês competência e class. cont.)
    df_despesas_provisionadas,
    df_despesas_rapidas,
    left_on=['ID Casa', 'Casa', 'Mes Competencia', 'Fornecedor', 'Classificação Contábil 1', 'Classificação Contábil 2'],
    right_on=['ID Casa', 'Casa', 'Mes Competencia', 'Fornecedor', 'Class_Cont_1', 'Class_Cont_2'],
    how='left'
  )

  df_provisionadas_nao_lancadas = df_merge[df_merge['ID_Despesa'].isna()] # Provisões não lançadas na T_DESPESA_RAPIDA
  # Nesse caso, Data Vencimento = Data Previsão
  df_valor_provisionadas_nao_lancadas = df_provisionadas_nao_lancadas.groupby(['Casa', 'Data Vencimento'], as_index=False)['Valor'].sum()
  df_valor_provisionadas_nao_lancadas.rename(columns={'Casa': 'Empresa', 'Data Vencimento': 'Data', 'Valor': 'Provisões Pendentes'}, inplace=True)
  
  # Prepara despesas não lançadas para exibição
  df_provisionadas_nao_lancadas = df_provisionadas_nao_lancadas[['ID Casa', 'Casa', 'Data Competência', 'Data Vencimento', 'Valor', 'Fornecedor', 'Descrição', 'Forma Pagamento', 'Classificação Contábil 1', 'Classificação Contábil 2']]
  
  return df_provisionadas_nao_lancadas, df_valor_provisionadas_nao_lancadas


# Função que organiza ordem e nomes das colunas dos dfs
def organiza_colunas(df, tipo, seletor_status_despesa=None):
  if tipo == 'bares agrupados' or tipo == 'projeção casa a casa':
    colunas = [  # colunas base
        "Data",
        "Saldo_Inicio_Dia",
        "Valor_Liquido_Recebido",
        "Valor_Projetado_Zig",
        "Receita_Projetada_Extraord",
        "Receita_Projetada_Eventos",
        "Despesas_Aprovadas_Pendentes",
        "Despesas_Pagas",
        "Saldo_Final",
    ]

    # adicionar provisionadas se necessário
    if seletor_status_despesa == "Todas Previstas":
        colunas.insert(-1, "Provisões Pendentes")
    
    # a casa importa nesse caso
    if tipo == 'projeção casa a casa':
      colunas.insert(0, "Empresa")

    df_organizado = df[colunas]

    # colunas para formatação (todas menos Data)
    colunas_formatar = [c for c in colunas if c != "Data"]

    df_organizado = format_columns_brazilian(
      df_organizado,
      colunas_formatar
    )

    # rename único
    df_organizado = df_organizado.rename(columns={
      "Empresa": "Casa",
      "Saldo_Inicio_Dia": "Saldo Início Dia",
      "Valor_Liquido_Recebido": "Valor Líquido Recebido",
      "Valor_Projetado_Zig": "Valor Zig Projetado",
      "Receita_Projetada_Extraord": "Receitas Extr Projetadas",
      "Receita_Projetada_Eventos": "Receitas Eventos Projetadas",
      "Despesas_Aprovadas_Pendentes": "Despesas Pendentes",
      "Despesas_Pagas": "Despesas Pagas",
      "Saldo_Final": "Saldo Final"
    })
  
  elif tipo =='despesas do dia':
    df_organizado = df[['Loja', 'Previsao_Pgto', 'Data_Vencimento', 'ID_Despesa', 'Status_Aprovacao_Diretoria', 'Parcelamento', 'ID_Parcela', 'Valor', 'Status_Pgto', 'Fornecedor']]
    df_organizado = df_organizado.rename(columns={
        "Loja": "Casa",
        "Previsao_Pgto": "Previsão Pgto",
        "Data_Vencimento": "Data Vencimento",
        "ID_Despesa": "ID Despesa",
        "Status_Aprovacao_Diretoria": "Status Aprovação Diretoria",
        "ID_Parcela": "ID Parcela",        
        "Status_Pgto": "Status Pgto",
    })

  elif tipo == 'receitas extraordinárias':
    df_organizado = df[['Empresa', 'ID_Receita_Extraordinária', 'Data_Vencimento_Parcela', 'Valor_Parcela', 'Classificação', 'Nome_Cliente', 'Observações']]
    df_organizado = df_organizado.rename(columns={
        "Empresa": "Casa",
        "ID_Receita_Extraordinária": "ID Receita Extraodinária",
        "Data_Vencimento_Parcela": "Data Vencimento Parcela",
        "Valor_Parcela": "Valor Parcela", 
        "Nome_Cliente": "Nome Cliente"       
    })

  elif tipo == 'receitas de eventos':
    df_organizado = df[['Empresa', 'ID_Evento', 'Data_Vencimento_Parcela', 'Valor_Parcela', 'Classificação', 'Nome_Cliente', 'Observações']]
    df_organizado = df_organizado.rename(columns={
        "Empresa": "Casa",
        "ID_Evento": "ID Evento",
        "Data_Vencimento_Parcela": "Data Vencimento Parcela",
        "Valor_Parcela": "Valor Parcela",     
        "Nome_Cliente": "Nome Cliente"   
    })

  return df_organizado


def exibe_provisoes_pendentes(df_provisoes_nao_lancadas, data_fim, bares_selecionados):
  with st.expander('Visualizar provisões pendentes/não lançadas'):
    df_provisoes_nao_lancadas_filtrado = df_provisoes_nao_lancadas[
      (df_provisoes_nao_lancadas['Casa'].isin(bares_selecionados)) &
      (df_provisoes_nao_lancadas['Data Vencimento'].dt.date <= data_fim)
    ].copy()

    if not df_provisoes_nao_lancadas_filtrado.empty:
        valor_total = df_provisoes_nao_lancadas_filtrado['Valor'].sum()
        df_provisoes_nao_lancadas_filtrado = format_columns_brazilian(df_provisoes_nao_lancadas_filtrado, ['Valor'])
        df_provisoes_nao_lancadas_filtrado.sort_values(by=['Casa', 'Data Vencimento'], inplace=True)
        st.dataframe(df_provisoes_nao_lancadas_filtrado, hide_index=True, width='stretch')
        st.write(f'**Total**: {format_brazilian(valor_total)}')
    else:
        st.success('Sem provisões pendentes!')


def config_feriados():
  calendario_brasil = Brazil()
  anos_desejados = list(range(2023, 2031))
  datas_feriados = []

  for ano in anos_desejados:
    feriados_ano = calendario_brasil.holidays(ano)
    datas_feriados.extend([feriado[0] for feriado in feriados_ano])

  serie_datas_feriados = pd.Series(datas_feriados, name='Data_Feriado')
  serie_datas_feriados = pd.to_datetime(serie_datas_feriados)
  nova_data = pd.to_datetime('2024-03-29')
  serie_datas_feriados = pd.concat([serie_datas_feriados, pd.Series([nova_data])]).reset_index(drop=True)

  return serie_datas_feriados


def calcular_taxa(row, taxas):
  if row['Tipo_Pagamento'] == 'DÉBITO':
    return row['Valor_Faturado'] * taxas['DÉBITO']
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 1:
    return row['Valor_Faturado'] * taxas['CRÉDITO_ANTECIPADO']
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 0:
    return row['Valor_Faturado'] * taxas['CRÉDITO_PADRAO']
  elif row['Tipo_Pagamento'] == 'APP':
    return row['Valor_Faturado'] * taxas['APP']
  elif row['Tipo_Pagamento'] == 'PIX':
    return row['Valor_Faturado'] * taxas['PIX']
  else:
    return 0.00


def ajustar_data_compensacao(row, serie_datas_feriados):
  if row['Tipo_Pagamento'] == 'DÉBITO':
    row['Data_Compensacao'] += pd.Timedelta(days=1)
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 1:
    row['Data_Compensacao'] += pd.Timedelta(days=1)
#########################     AQUI ESTÁ O CRÉDITO     ###############################
  elif row['Tipo_Pagamento'] == 'CRÉDITO' and row['Antecipacao_Credito'] == 0:
    row['Data_Compensacao'] += pd.Timedelta(days=31)
    ################################################
  elif row['Tipo_Pagamento'] in ['PIX', 'DINHEIRO', 'VOUCHER', 'ANTECIPADO']:
    row['Data_Compensacao'] += pd.Timedelta(days=1)
    #################### APP CONTA IGUAL CRÉDITO ########################
  elif row['Tipo_Pagamento'] == 'APP':
    row['Data_Compensacao'] += pd.Timedelta(days=31)

  if row['Data_Compensacao'] in serie_datas_feriados.values:
    row['Data_Compensacao'] += pd.Timedelta(days=1)

  if row['Data_Compensacao'].strftime('%A') == 'Sunday':
    row['Data_Compensacao'] += pd.Timedelta(days=1)
  elif row['Data_Compensacao'].strftime('%A') == 'Saturday':
    row['Data_Compensacao'] += pd.Timedelta(days=2)

  if row['Data_Compensacao'] in serie_datas_feriados.values:
    row['Data_Compensacao'] += pd.Timedelta(days=1)

  return row['Data_Compensacao']


def somar_total(df):
  colunas_numericas = df.select_dtypes(include=[int, float]).columns
  soma_colunas = df[colunas_numericas].sum().to_frame().T
  soma_colunas['Data'] = 'Total' 
  df_com_soma = pd.concat([df, soma_colunas], ignore_index=True)
  return df_com_soma


def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)